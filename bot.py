# =========================================================
# ZZHOMEY TELEGRAM BOT - SMART PRICE + LOG SYSTEM ADVANCED
# =========================================================
# FITUR TAMBAHAN:
# ✅ Log ke Excel
# ✅ Statistik pemakaian per distributor
# ✅ Auto rekap & export log bulanan otomatis
# =========================================================

import os
import time
import re
from datetime import datetime
from collections import Counter
from openpyxl import load_workbook, Workbook
from telegram import Update
from telegram.ext import ApplicationBuilder, MessageHandler, ContextTypes, filters

# ==============================
# KONFIGURASI
# ==============================
# TOKEN sekarang diambil dari Environment Variable agar aman saat upload ke GitHub
TOKEN = os.getenv("TOKEN")

if not TOKEN:
    raise ValueError("TOKEN tidak ditemukan. Set environment variable TOKEN terlebih dahulu.")

MASTER_FILE = "master.xlsx"
LOG_FILE = "usage_log.xlsx"


# ==============================
# DETEKSI HARGA
# ==============================
def detect_price_column(text):

    if not text:
        return "up 30%"

    text = text.lower()

    if "normal" in text:
        return "normal"

    match = re.search(r"(10|20|30|40|50)", text)
    if match:
        return f"up {match.group(1)}%"

    return "up 30%"


# ==============================
# AMBIL NOMOR HP
# ==============================
def extract_phone(text):
    if not text:
        return "Tidak dicantumkan"

    match = re.search(r"(08\d{8,13})", text)
    if match:
        return match.group(1)

    return "Tidak dicantumkan"


# ==============================
# SIMPAN LOG + STATISTIK
# ==============================
def save_log(username, fullname, phone, price):

    now = datetime.now()
    bulan = now.strftime("%Y-%m")

    if not os.path.exists(LOG_FILE):
        wb_log = Workbook()
        ws_log = wb_log.active
        ws_log.title = "Log"
        ws_log.append(["Tanggal", "Bulan", "Username", "Nama", "No HP", "Harga Dipakai"])
        wb_log.save(LOG_FILE)
        wb_log.close()

    wb_log = load_workbook(LOG_FILE)

    # ===== LOG DETAIL =====
    ws_log = wb_log["Log"]
    ws_log.append([
        now.strftime("%Y-%m-%d %H:%M:%S"),
        bulan,
        f"@{username}" if username else "Tidak ada username",
        fullname,
        phone,
        price.upper(),
    ])

    # ===== STATISTIK PER DISTRIBUTOR =====
    if "Statistik" not in wb_log.sheetnames:
        ws_stat = wb_log.create_sheet("Statistik")
        ws_stat.append(["Username", "Total Pemakaian"])
    else:
        ws_stat = wb_log["Statistik"]

    usage_counter = Counter()

    for row in ws_log.iter_rows(min_row=2, values_only=True):
        usage_counter[row[2]] += 1

    if ws_stat.max_row >= 2:
        ws_stat.delete_rows(2, ws_stat.max_row)

    for user, total in usage_counter.items():
        ws_stat.append([user, total])

    # ===== REKAP BULANAN =====
    if "Rekap Bulanan" not in wb_log.sheetnames:
        ws_month = wb_log.create_sheet("Rekap Bulanan")
        ws_month.append(["Bulan", "Total Pemakaian"])
    else:
        ws_month = wb_log["Rekap Bulanan"]

    monthly_counter = Counter()

    for row in ws_log.iter_rows(min_row=2, values_only=True):
        monthly_counter[row[1]] += 1

    if ws_month.max_row >= 2:
        ws_month.delete_rows(2, ws_month.max_row)

    for month, total in monthly_counter.items():
        ws_month.append([month, total])

    wb_log.save(LOG_FILE)
    wb_log.close()


# ==============================
# HANDLE FILE
# ==============================
async def handle_file(update: Update, context: ContextTypes.DEFAULT_TYPE):

    doc = update.message.document

    if not doc.file_name.endswith(".xlsx"):
        await update.message.reply_text("❌ Kirim file Excel (.xlsx)")
        return

    caption_text = update.message.caption or ""
    price_column = detect_price_column(caption_text)
    phone_number = extract_phone(caption_text)

    telegram_username = update.effective_user.username
    telegram_name = update.effective_user.full_name

    file = await doc.get_file()
    input_file = doc.file_name
    await file.download_to_drive(input_file)

    wb = None
    wb_master = None

    try:
        if not os.path.exists(MASTER_FILE):
            await update.message.reply_text("❌ File master.xlsx tidak ditemukan")
            return

        wb = load_workbook(input_file, data_only=False, keep_links=False)

        if "Template" not in wb.sheetnames:
            await update.message.reply_text("❌ Sheet 'Template' tidak ditemukan")
            return

        ws = wb["Template"]

        # Ambil header baris 3
        headers = []
        for col in range(1, ws.max_column + 1):
            headers.append(ws.cell(row=3, column=col).value)

        # Load master (formula mode)
        wb_master = load_workbook(MASTER_FILE, data_only=False)
        ws_master = wb_master.active

        master_columns_lower = {}
        for col in range(1, ws_master.max_column + 1):
            header_val = ws_master.cell(row=1, column=col).value
            if header_val is not None:
                master_columns_lower[str(header_val).strip().lower()] = col

        # Hapus data lama mulai baris 7
        if ws.max_row >= 7:
            ws.delete_rows(7, ws.max_row - 6)

        start_row = 7

        for i in range(2, ws_master.max_row + 1):
            for col_index, header in enumerate(headers, start=1):

                header_str = "" if header is None else str(header).strip().lower()
                target_cell = ws.cell(row=start_row + (i - 2), column=col_index)

                if header_str == "harga":
                    if price_column in master_columns_lower:
                        master_col_index = master_columns_lower[price_column]
                        target_cell.value = ws_master.cell(row=i, column=master_col_index).value
                    elif "up 30%" in master_columns_lower:
                        master_col_index = master_columns_lower["up 30%"]
                        target_cell.value = ws_master.cell(row=i, column=master_col_index).value
                    else:
                        target_cell.value = ""
                else:
                    if header_str in master_columns_lower:
                        master_col_index = master_columns_lower[header_str]
                        target_cell.value = ws_master.cell(row=i, column=master_col_index).value
                    else:
                        target_cell.value = ""

        # Tutup master setelah loop selesai
        if wb_master is not None:
            wb_master.close()
            wb_master = None

        wb.save(input_file)
        wb.close()
        wb = None

        # Simpan log + statistik
        save_log(telegram_username, telegram_name, phone_number, price_column)

        time.sleep(0.5)

        with open(input_file, "rb") as f:
            await update.message.reply_document(
                document=f,
                caption=(
                    f"Template sudah di sesuaikan dengan akun shopeemu menggunakan harga {price_column.upper()}. "
                    "Silahkan download file ini dan upload ke shopee."
                ),
            )

    except Exception as e:
        error_text = str(e)

        # Handle Protected View / invalid XML
        if "Unable to read workbook" in error_text or "invalid XML" in error_text:
            error_msg = """
❌ GAGAL MEMPROSES FILE

File Shopee masih dalam mode Protected View.
Silakan lakukan langkah berikut:
1️⃣ Buka file di Microsoft Excel
2️⃣ Klik tombol 'Enable Editing'
3️⃣ Klik Save / Simpan
4️⃣ Kirim ulang file tersebut ke bot

Setelah disimpan ulang, file akan bisa diproses.
"""
            await update.message.reply_text(error_msg.strip())
        else:
            await update.message.reply_text(f"❌ Error: {error_text}")

    finally:
        # Pastikan workbook tertutup sebelum hapus file (hindari file terkunci di Windows)
        try:
            if wb is not None:
                wb.close()
        except Exception:
            pass

        try:
            if wb_master is not None:
                wb_master.close()
        except Exception:
            pass

        # Tunggu sebentar agar handle file benar-benar release
        time.sleep(0.3)

        # Hapus file dari direktori baik sukses maupun gagal
        try:
            if os.path.exists(input_file):
                os.remove(input_file)
        except Exception:
            pass


# ==============================
# MAIN
# ==============================

def main():
    app = ApplicationBuilder().token(TOKEN).build()
    app.add_handler(MessageHandler(filters.Document.ALL, handle_file))

    print("🤖 Bot ZZHOMEY Aktif (ADVANCED LOG MODE)...")
    app.run_polling()


if __name__ == "__main__":
    main()
